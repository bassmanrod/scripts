#!/usr/bin/env python3
import requests
import json
import time
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
import pandas as pd

wb = Workbook()
worksheet = wb.active

url = 'https://xxxxxxxx.okta.com/api/v1/users'
headers = {'Accept': 'application/json'}
headers = {'Content-Type': 'application/json'}
headers = {'Authorization': 'xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx'}

req = requests.get(url, headers=headers)

parsed_json = json.loads(req.text)
#print(parsed_json)

worksheet.cell(row=1, column=1).value = 'Employee Name'
worksheet.cell(row=1, column=2).value = 'Status'
worksheet.cell(row=1, column=3).value = 'Reporting Manager'
worksheet.cell(row=1, column=4).value = 'Title'

worksheet.title = "Biotrack Okta Employee List"

for i in parsed_json:
    
    if not 'manager' in i['profile']:
        b = i['profile']['login']
        #print(b)
        status = i["status"]
        #print(status)
        
        if not 'title' in i['profile']:
            title = "no title assigned"
            manager = "no manager assigned"
            #print(manager)
            data = (b,status,manager,title)
            worksheet.append(data)
            print(data)
        
    else:
            
        b = i['profile']['login']
        #print(b)
        status = i["status"]
        #print(status)
        c = i['profile']['title']
        manager = i['profile']['manager']
        title = i['profile']['title']
        #print(manager)
        data = (b,status,manager,title)
        print(data)
        worksheet.append(data)
      
print("Done.")
wb.save('okta_biotrackthc.xlsx')




